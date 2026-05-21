"""导出选中角色到 Excel（图片从 Blob URL 读回，Excel 写到 Blob 返回 URL）。"""

from __future__ import annotations

import io
import time
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from .storage import blob_put, kv_get_json

EMBED_MAX_H_PX = 1024
EMBED_DISPLAY_H_PX = 480

HEADERS = [
    "#",
    "Original Name", "Original Gender", "Original Tags",
    "Original Tagline", "Original Opener",
    "New Name", "New Tags", "New Tagline (cha_set)", "New Opener (cha_set)",
    "Card Tagline", "Card Opener", "User Background",
    "Character Portrait", "User Portrait",
]


def _safe(s: Any) -> str:
    if s is None:
        return ""
    if isinstance(s, list):
        return ", ".join(map(str, s))
    return str(s)


def _resize_for_excel(img_bytes: bytes, max_h_px: int = EMBED_MAX_H_PX) -> bytes:
    with PILImage.open(io.BytesIO(img_bytes)) as im:
        im.load()
        w, h = im.size
        if h > max_h_px:
            ratio = max_h_px / h
            im = im.resize((int(w * ratio), int(h * ratio)), PILImage.LANCZOS)
        if im.mode not in ("RGB", "RGBA"):
            im = im.convert("RGB")
        buf = io.BytesIO()
        im.save(buf, format="PNG", optimize=True)
        buf.seek(0)
        return buf.getvalue()


def _fetch_blob(url: str) -> bytes | None:
    try:
        r = requests.get(url, timeout=30)
        r.raise_for_status()
        return r.content
    except Exception:
        return None


def collect_artifacts(cid: str) -> dict[str, Any]:
    enrich = kv_get_json(f"enriched:{cid}")
    char_p = kv_get_json(f"char_portraits:{cid}") or []
    user_p = kv_get_json(f"user_portraits:{cid}") or []
    return {
        "character_id": cid,
        "enrich": enrich,
        "char_portraits": char_p,
        "user_portraits": user_p,
    }


def list_candidates() -> list[dict[str, Any]]:
    items = kv_get_json("characters") or []
    out = []
    for ch in items:
        cid = ch.get("character_id")
        if not cid:
            continue
        art = collect_artifacts(cid)
        if not (art["enrich"] or art["char_portraits"] or art["user_portraits"]):
            continue
        out.append(
            {
                "character_id": cid,
                "name": ch.get("name", ""),
                "gender": ch.get("gender", ""),
                "image_url": ch.get("image_url") or ch.get("face_url") or "",
                "has_enrich": bool(art["enrich"]),
                "char_portrait_count": len(art["char_portraits"]),
                "user_portrait_count": len(art["user_portraits"]),
                "char_portrait_url": (art["char_portraits"][0]["image_url"]
                                       if art["char_portraits"] else None),
                "user_portrait_url": (art["user_portraits"][0]["image_url"]
                                       if art["user_portraits"] else None),
            }
        )
    out.sort(key=lambda x: (-int(x["has_enrich"]), -x["char_portrait_count"], x["name"]))
    return out


def export_to_excel(character_ids: list[str]) -> dict[str, Any]:
    items = kv_get_json("characters") or []
    index = {c.get("character_id"): c for c in items if c.get("character_id")}

    wb = Workbook()
    ws = wb.active
    ws.title = "Characters"

    header_font = Font(bold=True, color="FFFFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1f1f28")
    for col_i, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_i, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    widths = [4, 22, 12, 28, 48, 48, 22, 28, 56, 56, 48, 48, 32, 36, 36]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30

    row_h = 380

    for row_i, cid in enumerate(character_ids, start=2):
        ch = index.get(cid)
        if not ch:
            continue
        art = collect_artifacts(cid)
        enrich_result = (art.get("enrich") or {}).get("result") or {}
        cha_set = enrich_result.get("cha_set") or {}
        card = enrich_result.get("card") or {}
        user_set = enrich_result.get("user_set") or {}

        values = [
            row_i - 1,
            _safe(ch.get("name", "")),
            _safe(ch.get("gender", "")),
            _safe(ch.get("tags", [])),
            _safe(ch.get("introduction", "")),
            _safe(ch.get("greeting", "")),
            _safe(cha_set.get("name", "")),
            _safe(cha_set.get("core_tags", [])),
            _safe(cha_set.get("tagline", "")),
            _safe(cha_set.get("opener", "")),
            _safe(card.get("tagline", "")),
            _safe(card.get("opener", "")),
            _safe(user_set.get("background", "")),
            "", "",
        ]
        for col_i, v in enumerate(values, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row_i].height = row_h

        for col_letter, key, col_i in (
            ("N", "char_portraits", 14),
            ("O", "user_portraits", 15),
        ):
            paths = art.get(key) or []
            if not paths:
                continue
            url = paths[0].get("image_url")
            if not url:
                continue
            raw = _fetch_blob(url)
            if not raw:
                ws.cell(row=row_i, column=col_i, value="[image fetch failed]")
                continue
            try:
                img_bytes = _resize_for_excel(raw)
                xli = XLImage(io.BytesIO(img_bytes))
                xli.height = EMBED_DISPLAY_H_PX
                xli.width = int(EMBED_DISPLAY_H_PX * 9 / 16)
                ws.add_image(xli, f"{col_letter}{row_i}")
            except Exception as e:
                ws.cell(row=row_i, column=col_i, value=f"[image error] {e}")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    ts = time.strftime("%Y%m%d_%H%M%S")
    pathname = f"exports/characters_export_{ts}.xlsx"
    url = blob_put(
        pathname,
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return {"ok": True, "filename": f"characters_export_{ts}.xlsx", "url": url, "count": len(character_ids)}
